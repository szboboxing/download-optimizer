import os
import json
import base64
import threading
import tkinter as tk
from copy import copy
import requests
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

try:
    import openpyxl
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.formatting.rule import FormulaRule
    import shutil
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


BG_COLOR = "#F5F7FA"
CARD_COLOR = "#FFFFFF"
ACCENT = "#2196F3"
SUCCESS = "#4CAF50"
DANGER = "#F44336"
PURPLE = "#9C27B0"
TEXT_PRIMARY = "#212121"
TEXT_SECONDARY = "#666666"
TEXT_MUTED = "#9E9E9E"
BORDER = "#E0E0E0"
APP_VERSION = "3.11"

SPEC_HEADER_ROWS = 4
SPEC_DATA_START_ROW = SPEC_HEADER_ROWS + 1
SPEC_CHECKBOX_COLUMN = 11
SPEC_UPLOAD_STATUS_COLUMN = 12
SPEC_DUPLICATE_STATUS = "待上传（重复）"
SPEC_DEFAULT_LEFT_COLUMN = "S"
SPEC_DEFAULT_RIGHT_COLUMN = "A"
SPEC_COLUMN_CHOICES = tuple(
    chr(column) for column in range(ord("A"), ord("Z") + 1)
) + tuple(
    f"A{chr(column)}" for column in range(ord("A"), ord("Z") + 1)
)


def _snapshot_worksheet_row(ws, row_number, max_column):
    cells = []
    for column in range(1, max_column + 1):
        cell = ws.cell(row=row_number, column=column)
        cells.append({
            "value": cell.value,
            "style": copy(cell._style),
            "hyperlink": copy(cell.hyperlink),
            "comment": copy(cell.comment),
        })

    row_dimension = None
    if row_number in ws.row_dimensions:
        row_dimension = copy(ws.row_dimensions[row_number])
    return cells, row_dimension


def _restore_worksheet_row(ws, row_number, snapshot):
    cells, row_dimension = snapshot
    for column, cell_data in enumerate(cells, start=1):
        cell = ws.cell(row=row_number, column=column)
        if isinstance(cell, MergedCell):
            continue
        cell.value = cell_data["value"]
        cell._style = copy(cell_data["style"])
        cell.hyperlink = copy(cell_data["hyperlink"])
        cell.comment = copy(cell_data["comment"])

    if row_dimension is None:
        ws.row_dimensions.pop(row_number, None)
    else:
        restored_dimension = copy(row_dimension)
        restored_dimension.index = row_number
        ws.row_dimensions[row_number] = restored_dimension


def _reorder_worksheet_rows(ws, ordered_rows, start_row, max_column):
    snapshots = {
        row_number: _snapshot_worksheet_row(ws, row_number, max_column)
        for row_number in ordered_rows
    }
    for target_row, source_row in enumerate(ordered_rows, start=start_row):
        _restore_worksheet_row(ws, target_row, snapshots[source_row])


def _normalize_spec_compare_columns(left_column, right_column):
    def normalize(value, field_name):
        label = str(value or "").strip().upper()
        if (
            not label
            or len(label) > 3
            or any(character < "A" or character > "Z" for character in label)
        ):
            raise ValueError(
                f"{field_name}“{value}”无效，请输入 A 至 XFD 范围内的列字母"
            )
        try:
            index = column_index_from_string(label)
        except ValueError as error:
            raise ValueError(f"{field_name}“{value}”无效") from error
        if index > 16384:
            raise ValueError(f"{field_name}“{label}”超出 XFD 列")
        return label, index

    left_label, left_index = normalize(
        left_column, "横杠左边对应列"
    )
    right_label, right_index = normalize(
        right_column, "横杠右边对应列"
    )
    return left_label, left_index, right_label, right_index


def _validate_spec_compare_columns(excel_data, compare_columns):
    max_column = max((len(row) for row in excel_data), default=0)
    missing_columns = [
        label
        for label, index in (
            (compare_columns[0], compare_columns[1]),
            (compare_columns[2], compare_columns[3]),
        )
        if index > max_column
    ]
    if missing_columns:
        last_column = get_column_letter(max_column) if max_column else "无"
        raise ValueError(
            f"表格最大列为 {last_column}，不存在所选列："
            f"{'、'.join(missing_columns)}"
        )


def _spec_row_value(row, column_index):
    if column_index <= 0 or column_index > len(row):
        return ""
    return str(row[column_index - 1]).strip()


def _spec_status_is_marked(checkbox_value, upload_status):
    return (
        checkbox_value not in (None, "")
        or upload_status not in (None, "")
    )


def _validate_spec_status_columns(excel_data):
    max_column = max((len(row) for row in excel_data), default=0)
    if max_column < SPEC_UPLOAD_STATUS_COLUMN:
        raise ValueError(
            "表格中不存在现有 K/L 状态列，已停止处理且不会新增列"
        )

    header_rows = excel_data[:SPEC_HEADER_ROWS]
    checkbox_headers = {
        _spec_row_value(row, SPEC_CHECKBOX_COLUMN)
        for row in header_rows
    }
    upload_headers = {
        _spec_row_value(row, SPEC_UPLOAD_STATUS_COLUMN)
        for row in header_rows
    }
    if "复核" not in checkbox_headers or "需上传" not in upload_headers:
        raise ValueError(
            "未识别到现有 K 列“复核”和 L 列“需上传”表头。"
            "为避免覆盖原数据，已停止处理且不会新增列"
        )


def _spec_status_formula(row_number):
    return f'=IF(K{row_number}=TRUE,"已完成上传","待上传")'


def _is_spec_status_formula(value):
    return (
        isinstance(value, str)
        and value.startswith("=IF(")
        and '"已完成上传"' in value
        and '"待上传"' in value
    )


def _set_worksheet_columns_hidden(ws, start_column, end_column):
    for column in range(start_column, end_column + 1):
        dimension = ws.column_dimensions[get_column_letter(column)]
        dimension.hidden = True
        dimension.outlineLevel = max(dimension.outlineLevel or 0, 1)
    ws.column_dimensions[
        get_column_letter(end_column + 1)
    ].collapsed = True


def _worksheet_columns_are_hidden(ws, start_column, end_column):
    for column in range(start_column, end_column + 1):
        if not any(
            dimension.hidden
            and (
                dimension.min
                if dimension.min is not None
                else column_index_from_string(key)
            ) <= column
            and (
                dimension.max
                if dimension.max is not None
                else column_index_from_string(key)
            ) >= column
            for key, dimension in ws.column_dimensions.items()
        ):
            return False
    return True


def _workbook_columns_are_hidden(file_path, start_column, end_column):
    wb = openpyxl.load_workbook(file_path, read_only=False)
    try:
        return _worksheet_columns_are_hidden(
            wb.active, start_column, end_column
        )
    finally:
        wb.close()


AI_PROVIDERS = {
    "硅基流动": {
        "base_url": "https://api.siliconflow.cn/v1",
        "models": [
            "Qwen/Qwen2.5-7B-Instruct",
            "Qwen/Qwen2.5-14B-Instruct",
            "Qwen/Qwen2.5-32B-Instruct",
            "Qwen/Qwen2.5-72B-Instruct",
            "deepseek-ai/DeepSeek-V3.2",
            "deepseek-ai/DeepSeek-V4-Pro",
            "deepseek-ai/DeepSeek-R1",
            "Pro/zai-org/GLM-5.1",
            "Pro/moonshotai/Kimi-K2.6",
        ],
        "default_model": "Qwen/Qwen2.5-7B-Instruct",
    },
    "DeepSeek": {
        "base_url": "https://api.deepseek.com/v1",
        "models": ["deepseek-chat", "deepseek-reasoner"],
        "default_model": "deepseek-chat",
    },
    "豆包 (火山引擎)": {
        "base_url": "https://ark.cn-beijing.volces.com/api/v3",
        "models": ["doubao-pro-32k", "doubao-pro-128k", "doubao-lite-32k", "doubao-1.5-pro"],
        "default_model": "doubao-pro-32k",
    },
    "元宝 (百度千帆)": {
        "base_url": "https://qianfan.baidubce.com/v2",
        "models": ["ernie-4.0-turbo-8k", "ernie-4.0-turbo-128k", "ernie-4.0", "qianfan-max"],
        "default_model": "ernie-4.0-turbo-8k",
    },
}


class MainApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"下载流程优化工具 v{APP_VERSION}")
        self.root.geometry("860x650")
        self.root.configure(bg=BG_COLOR)
        self.root.resizable(False, False)
        self._setup_styles()
        self._build_ui()

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Accent.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground="white", background=ACCENT,
                        borderwidth=0, padding=(12, 6))
        style.map("Accent.TButton", background=[("active", "#1976D2")])
        style.configure("Success.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground="white", background=SUCCESS,
                        borderwidth=0, padding=(12, 6))
        style.map("Success.TButton", background=[("active", "#388E3C")])
        style.configure("Danger.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground=DANGER, background=CARD_COLOR,
                        borderwidth=1, relief="solid", padding=(12, 6))
        style.map("Danger.TButton", background=[("active", "#FFEBEE")])
        style.configure("Ghost.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground=TEXT_PRIMARY, background=CARD_COLOR,
                        borderwidth=1, relief="solid", padding=(12, 6))
        style.map("Ghost.TButton", background=[("active", "#F5F5F5")])

    def _build_ui(self):
        container = tk.Frame(self.root, bg=BG_COLOR)
        container.pack(fill="both", expand=True, padx=20, pady=20)

        tk.Label(container, text="下载流程优化工具",
                 font=("Microsoft YaHei", 22, "bold"),
                 fg=ACCENT, bg=BG_COLOR).pack(pady=(10, 4))
        tk.Label(container, text=f"v{APP_VERSION}  —  选择一种模式开始",
                 font=("Microsoft YaHei", 10),
                 fg=TEXT_SECONDARY, bg=BG_COLOR).pack(pady=(0, 20))

        cards_frame = tk.Frame(container, bg=BG_COLOR)
        cards_frame.pack(pady=10)
        cards_frame.columnconfigure(0, weight=1, minsize=360)
        cards_frame.columnconfigure(1, weight=1, minsize=360)

        self._create_mode_card(cards_frame, "📁", "批量重命名文件夹",
                                "删除前N位字符 或 后加内容", None,
                                ACCENT, self._open_folder_rename, 0, 0)
        self._create_mode_card(cards_frame, "📄", "批量重命名文件",
                                "添加日期、版本号、自定义字符", None,
                                SUCCESS, self._open_file_rename, 0, 1)
        self._create_mode_card(cards_frame, "🤖", "AI 智能助手",
                                "对话式 AI，支持多平台切换", None,
                                PURPLE, self._open_ai_chat, 1, 0)
        self._create_mode_card(cards_frame, "__shield_check__", "规约上传数据表准备",
                                "比对生成报网公司结算数据表格",
                                "上传第三方填写审定信息",
                                "#FF9800", self._open_spec_upload, 1, 1)

    def _create_shield_check_icon(self, parent):
        c = tk.Canvas(parent, width=70, height=65,
                      bg=CARD_COLOR, highlightthickness=0, cursor="hand2")
        c.create_oval(8, 2, 62, 56, fill="#FFF3E0", outline="#FF9800", width=2)
        c.create_oval(14, 8, 56, 50, fill="#FFB74D", outline="#F57C00", width=1)
        c.create_oval(20, 14, 50, 44, fill="#FF9800", outline="")
        c.create_line(25, 30, 33, 38, fill="white", width=4, capstyle="round")
        c.create_line(33, 38, 46, 22, fill="white", width=4, capstyle="round")
        c.create_oval(28, 17, 32, 21, fill="#FFD700", outline="")
        c.create_oval(34, 15, 38, 19, fill="#FFD700", outline="")
        c.create_oval(40, 17, 44, 21, fill="#FFD700", outline="")
        c.create_oval(25, 42, 45, 60, fill="#FFE0B2", outline="#FFB74D", width=2)
        c.create_polygon(35, 46, 30, 52, 40, 52, fill="#F57C00", outline="")
        return c

    def _create_mode_card(
            self, parent, icon, title, desc, desc2, color, cmd, row, col,
            columnspan=1):
        card = tk.Frame(parent, bg=CARD_COLOR, cursor="hand2",
                        highlightbackground=BORDER,
                        highlightthickness=1)
        card.grid(
            row=row, column=col, columnspan=columnspan,
            padx=12, pady=8, sticky="nsew"
        )

        inner = tk.Frame(card, bg=CARD_COLOR)
        inner.pack(padx=24, pady=28)

        if isinstance(icon, str) and icon.startswith("__"):
            if icon == "__shield_check__":
                icon_widget = self._create_shield_check_icon(inner)
            else:
                icon_widget = self._create_shield_check_icon(inner)
            icon_widget.pack()
        elif isinstance(icon, tk.Widget):
            icon_widget = icon
            icon_widget.pack()
        else:
            icon_widget = tk.Label(inner, text=icon, font=("Segoe UI Emoji", 42),
                                  bg=CARD_COLOR, fg=color)
            icon_widget.pack()
        tk.Label(inner, text=title, font=("Microsoft YaHei", 14, "bold"),
                 bg=CARD_COLOR, fg=TEXT_PRIMARY).pack(pady=(8, 4))
        if desc2:
            tk.Label(inner, text=desc2, font=("Microsoft YaHei", 8),
                     bg=CARD_COLOR, fg=color,
                     wraplength=200, justify="center").pack()
        tk.Label(inner, text=desc, font=("Microsoft YaHei", 9),
                 bg=CARD_COLOR, fg=TEXT_SECONDARY,
                 wraplength=200, justify="center").pack()

        card.bind("<Button-1>", lambda e: cmd())
        inner.bind("<Button-1>", lambda e: cmd())
        icon_widget.bind("<Button-1>", lambda e: cmd())
        for w in inner.winfo_children():
            w.bind("<Button-1>", lambda e: cmd())

    def _open_folder_rename(self):
        self.root.destroy()
        root = tk.Tk()
        FolderRenameApp(root)
        root.mainloop()

    def _open_file_rename(self):
        self.root.destroy()
        root = tk.Tk()
        FileRenameApp(root)
        root.mainloop()

    def _open_ai_chat(self):
        try:
            self.root.destroy()
            root = tk.Tk()
            AIChatApp(root)
            root.mainloop()
        except Exception as e:
            import traceback
            traceback.print_exc()
            try:
                messagebox.showerror("启动错误", f"AI助手启动失败：\n{e}")
            except Exception:
                pass

    def _open_spec_upload(self):
        if not HAS_OPENPYXL:
            messagebox.showerror("依赖缺失",
                "缺少 openpyxl 库。\n请运行：pip install openpyxl")
            return
        try:
            self.root.destroy()
            root = tk.Tk()
            SpecUploadApp(root)
            root.mainloop()
        except Exception as e:
            import traceback
            traceback.print_exc()
            try:
                messagebox.showerror("启动错误", f"规约上传数据表准备启动失败：\n{e}")
            except Exception:
                pass

class FolderRenameApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"批量文件夹重命名工具 v{APP_VERSION}")
        self.root.geometry("700x680")
        self.root.configure(bg=BG_COLOR)
        self.root.resizable(False, False)

        self.folder_path = tk.StringVar()
        self.status_text = tk.StringVar(value="请选择要处理的文件夹")
        self.progress_var = tk.DoubleVar(value=0.0)
        self.remove_chars = tk.IntVar(value=3)
        self.rename_mode = tk.StringVar(value="remove")
        self.suffix_text = tk.StringVar(value="-重复送审项目")
        self.prefix_text = tk.StringVar(value=datetime.now().strftime("%Y%m%d_"))

        self.rename_history = []
        self.is_processing = False

        self._setup_styles()
        self._build_ui()

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Card.TLabelframe",
                        background=CARD_COLOR, foreground=TEXT_PRIMARY,
                        bordercolor=BORDER)
        style.configure("Card.TLabelframe.Label",
                        font=("Microsoft YaHei", 10, "bold"),
                        foreground=ACCENT, background=CARD_COLOR)
        style.configure("Accent.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground="white", background=ACCENT,
                        borderwidth=0, padding=(12, 8))
        style.map("Accent.TButton", background=[("active", "#1976D2")])
        style.configure("Success.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground="white", background=SUCCESS,
                        borderwidth=0, padding=(12, 8))
        style.map("Success.TButton", background=[("active", "#388E3C")])
        style.configure("Danger.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground=DANGER, background=CARD_COLOR,
                        borderwidth=1, relief="solid", padding=(12, 8))
        style.map("Danger.TButton", background=[("active", "#FFEBEE")])
        style.configure("Ghost.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground=TEXT_PRIMARY, background=CARD_COLOR,
                        borderwidth=1, relief="solid", padding=(12, 8))
        style.map("Ghost.TButton", background=[("active", "#F5F5F5")])
        style.configure("Horizontal.TProgressbar",
                        troughcolor=BORDER, background=SUCCESS,
                        thickness=8, borderwidth=0)

    def _build_ui(self):
        container = tk.Frame(self.root, bg=BG_COLOR)
        container.pack(fill="both", expand=True, padx=16, pady=12)

        header = tk.Frame(container, bg=BG_COLOR)
        header.pack(fill="x", pady=(0, 8))
        ttk.Label(header, text="📁 批量文件夹重命名工具",
                  font=("Microsoft YaHei", 16, "bold"),
                  foreground=ACCENT, background=BG_COLOR).pack(side="left")
        back_btn = tk.Label(header, text="← 返回主菜单",
                            font=("Microsoft YaHei", 9),
                            fg=TEXT_SECONDARY, bg=BG_COLOR,
                            cursor="hand2")
        back_btn.pack(side="right")
        back_btn.bind("<Button-1>", lambda e: self._go_main())

        sep = tk.Frame(container, bg=BORDER, height=1)
        sep.pack(fill="x", pady=(0, 10))

        rule_card = ttk.LabelFrame(container, text=" 重命名规则 ",
                                    style="Card.TLabelframe", padding=10)
        rule_card.pack(fill="x", pady=(0, 8))
        rule_inner = tk.Frame(rule_card, bg=CARD_COLOR)
        rule_inner.pack(fill="x")

        tk.Label(rule_inner, text="模式：",
                 font=("Microsoft YaHei", 10, "bold"),
                 bg=CARD_COLOR, fg=TEXT_PRIMARY).grid(row=0, column=0, sticky="w")

        mode_frame = tk.Frame(rule_inner, bg=CARD_COLOR)
        mode_frame.grid(row=0, column=1, columnspan=6, sticky="w", pady=(0, 6))

        ttk.Radiobutton(mode_frame, text="删除前N位字符",
                         variable=self.rename_mode, value="remove",
                         command=self._on_mode_change).pack(side="left", padx=(0, 16))
        ttk.Radiobutton(mode_frame, text="文件夹改名（后加内容）",
                         variable=self.rename_mode, value="suffix",
                         command=self._on_mode_change).pack(side="left", padx=(0, 16))
        ttk.Radiobutton(mode_frame, text="文件夹改名（前加内容）",
                         variable=self.rename_mode, value="prefix",
                         command=self._on_mode_change).pack(side="left")

        self.param_frame = tk.Frame(rule_inner, bg=CARD_COLOR)
        self.param_frame.grid(row=1, column=0, columnspan=7, sticky="w", pady=(4, 0))
        self._build_param_controls()

        path_card = ttk.LabelFrame(container, text=" 文件夹路径 ",
                                    style="Card.TLabelframe", padding=10)
        path_card.pack(fill="x", pady=(0, 8))
        path_inner = tk.Frame(path_card, bg=CARD_COLOR)
        path_inner.pack(fill="x")
        tk.Entry(path_inner, textvariable=self.folder_path,
                  font=("Microsoft YaHei", 10), relief="solid", bd=1).pack(side="left", fill="x", expand=True, padx=(0, 8), pady=6)
        ttk.Button(path_inner, text="浏览...",
                    style="Accent.TButton",
                    command=self._browse_folder, width=8).pack(side="right", pady=6)

        action_row = tk.Frame(container, bg=BG_COLOR)
        action_row.pack(fill="x", pady=(0, 8))
        action_row.columnconfigure(0, weight=1)
        action_row.columnconfigure(1, weight=1)
        action_row.columnconfigure(2, weight=1)

        self.start_btn = ttk.Button(action_row, text="▶  开始批量重命名",
                                     style="Success.TButton",
                                     command=self._start_rename)
        self.start_btn.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        preview_btn = ttk.Button(action_row, text="👁  预览",
                                  style="Ghost.TButton",
                                  command=self._preview_rename)
        preview_btn.grid(row=0, column=1, sticky="ew", padx=6)
        self.undo_btn = ttk.Button(action_row, text="↩  后悔，回退",
                                    style="Danger.TButton",
                                    command=self._undo_rename, state="disabled")
        self.undo_btn.grid(row=0, column=2, sticky="ew", padx=(6, 0))

        progress_card = ttk.LabelFrame(container, text=" 处理进度 ",
                                        style="Card.TLabelframe", padding=10)
        progress_card.pack(fill="x", pady=(0, 8))
        self.progress_bar = ttk.Progressbar(progress_card, variable=self.progress_var,
                                             maximum=100, style="Horizontal.TProgressbar")
        self.progress_bar.pack(fill="x", pady=(0, 6))
        tk.Label(progress_card, textvariable=self.status_text,
                  font=("Microsoft YaHei", 9),
                  fg=TEXT_SECONDARY, bg=CARD_COLOR).pack(anchor="w")

        log_card = ttk.LabelFrame(container, text=" 预览 / 日志 ",
                                    style="Card.TLabelframe", padding=10)
        log_card.pack(fill="both", expand=True)
        self.log_text = tk.Text(log_card, height=15, font=("Consolas", 9),
                                 state="disabled", relief="flat",
                                 bg="#FAFAFA", fg=TEXT_PRIMARY,
                                 selectbackground=ACCENT, selectforeground="white",
                                 wrap="word", padx=8, pady=6)
        self.log_text.pack(fill="both", expand=True)

    def _log(self, message):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{ts}] {message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self.root.update_idletasks()

    def _clear_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    def _go_main(self):
        self.root.destroy()
        main = tk.Tk()
        MainApp(main)
        main.mainloop()

    def _build_param_controls(self):
        for w in self.param_frame.winfo_children():
            w.destroy()
        mode = self.rename_mode.get()
        if mode == "remove":
            tk.Label(self.param_frame, text="删除前",
                     font=("Microsoft YaHei", 10),
                     bg=CARD_COLOR, fg=TEXT_PRIMARY).grid(row=0, column=0, sticky="w")
            self.remove_combo = ttk.Combobox(self.param_frame,
                                               values=[str(i) for i in range(1, 9)],
                                               width=3, state="readonly",
                                               font=("Microsoft YaHei", 10))
            self.remove_combo.set(str(self.remove_chars.get()))
            self.remove_combo.grid(row=0, column=1, sticky="w", padx=(4, 4))
            self.remove_combo.bind("<<ComboboxSelected>>", self._on_rule_change)
            tk.Label(self.param_frame, text="位字符  (1-8位可选)",
                     font=("Microsoft YaHei", 9),
                     bg=CARD_COLOR, fg=TEXT_PRIMARY).grid(row=0, column=2, sticky="w")
        elif mode == "suffix":
            tk.Label(self.param_frame, text="添加内容：",
                     font=("Microsoft YaHei", 10),
                     bg=CARD_COLOR, fg=TEXT_PRIMARY).grid(row=0, column=0, sticky="w")
            suffix_entry = tk.Entry(self.param_frame, textvariable=self.suffix_text,
                                     font=("Microsoft YaHei", 10), width=25,
                                     relief="solid", bd=1)
            suffix_entry.grid(row=0, column=1, sticky="w", padx=(4, 0))
            suffix_entry.bind("<KeyRelease>", lambda e: self._auto_preview())
            tk.Label(self.param_frame, text="  将加在文件夹名后面",
                     font=("Microsoft YaHei", 8),
                     bg=CARD_COLOR, fg=TEXT_MUTED).grid(row=0, column=2, sticky="w")
        elif mode == "prefix":
            tk.Label(self.param_frame, text="添加内容：",
                     font=("Microsoft YaHei", 10),
                     bg=CARD_COLOR, fg=TEXT_PRIMARY).grid(row=0, column=0, sticky="w")
            prefix_entry = tk.Entry(self.param_frame, textvariable=self.prefix_text,
                                     font=("Microsoft YaHei", 10), width=25,
                                     relief="solid", bd=1)
            prefix_entry.grid(row=0, column=1, sticky="w", padx=(4, 0))
            prefix_entry.bind("<KeyRelease>", lambda e: self._auto_preview())
            tk.Label(self.param_frame, text="  将加在文件夹名前面",
                     font=("Microsoft YaHei", 8),
                     bg=CARD_COLOR, fg=TEXT_MUTED).grid(row=0, column=2, sticky="w")

    def _on_mode_change(self):
        self._build_param_controls()
        self._auto_preview()

    def _on_rule_change(self, event=None):
        new_val = int(self.remove_combo.get())
        self.remove_chars.set(new_val)
        self._auto_preview()

    def _browse_folder(self):
        folder = filedialog.askdirectory(title="请选择要处理的文件夹")
        if folder:
            self.folder_path.set(folder)
            self.status_text.set(f"已选择：{folder}")
            self._auto_preview()

    def _get_subfolders(self, path):
        try:
            items = os.listdir(path)
            subfolders = [i for i in items if os.path.isdir(os.path.join(path, i))]
            return subfolders
        except Exception as e:
            messagebox.showerror("错误", f"无法读取文件夹：{e}")
            return []

    def _compute_new_name(self, name):
        mode = self.rename_mode.get()
        if mode == "remove":
            remove_n = self.remove_chars.get()
            if len(name) <= remove_n:
                return None
            return name[remove_n:]
        elif mode == "suffix":
            suffix = self.suffix_text.get()
            if not suffix:
                return None
            return name + suffix
        elif mode == "prefix":
            prefix = self.prefix_text.get()
            if not prefix:
                return None
            return prefix + name
        return None

    def _auto_preview(self):
        folder = self.folder_path.get().strip()
        if not folder or not os.path.isdir(folder):
            return
        self._clear_log()
        subfolders = self._get_subfolders(folder)
        if not subfolders:
            self._log("[预览] 当前文件夹下没有子文件夹。")
            return
        rename_list = []
        skip_list = []
        conflict_list = []
        for name in subfolders:
            new_name = self._compute_new_name(name)
            if new_name is None:
                skip_list.append(name)
            else:
                new_path = os.path.join(folder, new_name)
                if os.path.exists(new_path) and new_path != os.path.join(folder, name):
                    conflict_list.append((name, new_name))
                rename_list.append((name, new_name))
        self._log(f"[预览] 共 {len(subfolders)} 个子文件夹：")
        self._log(f"  待重命名：{len(rename_list)} 个，跳过：{len(skip_list)} 个")
        self._log("")
        for old, new in rename_list:
            flag = " ⚠冲突" if (old, new) in conflict_list else ""
            self._log(f"  {old}  →  {new}{flag}")
        if skip_list:
            self._log("")
            for name in skip_list:
                self._log(f"  {name}  →  [跳过]")
        self._log(f"[预览] 可重命名 {len(rename_list) - len(conflict_list)} 个")

    def _preview_rename(self):
        folder = self.folder_path.get().strip()
        if not folder:
            messagebox.showwarning("警告", "请先选择一个文件夹！")
            return
        self._auto_preview()

    def _start_rename(self):
        folder = self.folder_path.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("警告", "请先选择有效的文件夹！")
            return
        mode = self.rename_mode.get()
        if mode == "suffix" and not self.suffix_text.get():
            messagebox.showwarning("警告", "请填写要添加的内容！")
            return
        subfolders = self._get_subfolders(folder)
        if not subfolders:
            messagebox.showinfo("提示", "没有子文件夹。")
            return
        self._clear_log()
        self.rename_history = []
        self.is_processing = True
        self.start_btn.configure(state="disabled")
        self.undo_btn.configure(state="disabled")
        threading.Thread(target=self._do_rename,
                         args=(folder, subfolders), daemon=True).start()

    def _do_rename(self, folder, subfolders):
        success_count = fail_count = skip_count = conflict_count = 0
        total = len(subfolders)
        start_time = datetime.now()
        mode = self.rename_mode.get()
        self.progress_var.set(0)
        self.status_text.set("正在处理...")

        for i, name in enumerate(subfolders):
            old_path = os.path.join(folder, name)
            new_name = self._compute_new_name(name)
            if new_name is None:
                skip_count += 1
            else:
                new_path = os.path.join(folder, new_name)
                if os.path.exists(new_path) and new_path != old_path:
                    conflict_count += 1
                    fail_count += 1
                else:
                    try:
                        os.rename(old_path, new_path)
                        success_count += 1
                        self.rename_history.append((folder, name, new_name))
                    except Exception as e:
                        fail_count += 1
            progress = (i + 1) / total * 100
            self.progress_var.set(progress)
            self.status_text.set(f"处理中... {i+1}/{total}")
            self.root.update_idletasks()

        duration = (datetime.now() - start_time).total_seconds()
        summary = f"完成！成功 {success_count}，失败 {fail_count}，跳过 {skip_count}"
        self.status_text.set(summary)
        self.progress_var.set(100)
        self.start_btn.configure(state="normal")
        self.is_processing = False
        if success_count > 0 and self.rename_history:
            self.undo_btn.configure(state="normal")
        messagebox.showinfo("完成",
            f"成功：{success_count}\n失败：{fail_count}\n跳过：{skip_count}\n耗时：{duration:.2f}s")

    def _undo_rename(self):
        if self.is_processing:
            return
        if not self.rename_history:
            messagebox.showinfo("提示", "没有可回退的操作。")
            return
        count = len(self.rename_history)
        if not messagebox.askyesno("确认",
            f"回退最近的 {count} 个重命名操作？"):
            return
        self.start_btn.configure(state="disabled")
        self.undo_btn.configure(state="disabled")
        undo_history = list(reversed(self.rename_history))
        self.rename_history = []
        self.is_processing = True
        success_count = fail_count = 0
        total = len(undo_history)
        self.progress_var.set(0)
        self.status_text.set("正在回退...")
        for i, (f, orig, curr) in enumerate(undo_history):
            curr_path = os.path.join(f, curr)
            orig_path = os.path.join(f, orig)
            if not os.path.exists(curr_path):
                fail_count += 1
            elif os.path.exists(orig_path) and orig_path != curr_path:
                fail_count += 1
            else:
                try:
                    os.rename(curr_path, orig_path)
                    success_count += 1
                except Exception:
                    fail_count += 1
            progress = (i + 1) / total * 100
            self.progress_var.set(progress)
            self.status_text.set(f"回退中... {i+1}/{total}")
            self.root.update_idletasks()
        summary = f"回退完成！成功 {success_count}，失败 {fail_count}"
        self.status_text.set(summary)
        self.progress_var.set(100)
        self.start_btn.configure(state="normal")
        self.is_processing = False
        messagebox.showinfo("回退完成", summary)


class FileRenameApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"批量文件重命名工具 v{APP_VERSION}")
        self.root.geometry("700x680")
        self.root.configure(bg=BG_COLOR)
        self.root.resizable(False, False)

        self.folder_path = tk.StringVar()
        self.status_text = tk.StringVar(value="请选择要处理的文件夹")
        self.progress_var = tk.DoubleVar(value=0.0)

        self.rename_action = tk.StringVar(value="date")
        self.date_format = tk.StringVar(value="%Y%m%d")
        self.version_text = tk.StringVar(value="v1.0")
        self.custom_chars = tk.StringVar(value="前缀")
        self.add_position = tk.StringVar(value="prefix")
        self.custom_count = tk.IntVar(value=1)

        self.rename_history = []
        self.is_processing = False

        self._setup_styles()
        self._build_ui()

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Card.TLabelframe",
                        background=CARD_COLOR, foreground=TEXT_PRIMARY,
                        bordercolor=BORDER)
        style.configure("Card.TLabelframe.Label",
                        font=("Microsoft YaHei", 10, "bold"),
                        foreground=SUCCESS, background=CARD_COLOR)
        style.configure("Accent.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground="white", background=ACCENT,
                        borderwidth=0, padding=(12, 8))
        style.configure("Success.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground="white", background=SUCCESS,
                        borderwidth=0, padding=(12, 8))
        style.map("Success.TButton", background=[("active", "#388E3C")])
        style.configure("Danger.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground=DANGER, background=CARD_COLOR,
                        borderwidth=1, relief="solid", padding=(12, 8))
        style.configure("Ghost.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground=TEXT_PRIMARY, background=CARD_COLOR,
                        borderwidth=1, relief="solid", padding=(12, 8))
        style.configure("Horizontal.TProgressbar",
                        troughcolor=BORDER, background=SUCCESS,
                        thickness=8, borderwidth=0)

    def _build_ui(self):
        container = tk.Frame(self.root, bg=BG_COLOR)
        container.pack(fill="both", expand=True, padx=16, pady=12)

        header = tk.Frame(container, bg=BG_COLOR)
        header.pack(fill="x", pady=(0, 8))
        ttk.Label(header, text="📄 批量文件重命名工具",
                  font=("Microsoft YaHei", 16, "bold"),
                  foreground=SUCCESS, background=BG_COLOR).pack(side="left")
        back_btn = tk.Label(header, text="← 返回主菜单",
                            font=("Microsoft YaHei", 9),
                            fg=TEXT_SECONDARY, bg=BG_COLOR,
                            cursor="hand2")
        back_btn.pack(side="right")
        back_btn.bind("<Button-1>", lambda e: self._go_main())

        sep = tk.Frame(container, bg=BORDER, height=1)
        sep.pack(fill="x", pady=(0, 10))

        rule_card = ttk.LabelFrame(container, text=" 重命名规则 ",
                                    style="Card.TLabelframe", padding=10)
        rule_card.pack(fill="x", pady=(0, 8))

        action_frame = tk.Frame(rule_card, bg=CARD_COLOR)
        action_frame.pack(fill="x", pady=(0, 8))
        tk.Label(action_frame, text="操作类型：",
                 font=("Microsoft YaHei", 10, "bold"),
                 bg=CARD_COLOR, fg=TEXT_PRIMARY).pack(side="left")

        ttk.Radiobutton(action_frame, text="添加当前日期",
                         variable=self.rename_action, value="date",
                         command=self._on_action_change).pack(side="left", padx=(8, 12))
        ttk.Radiobutton(action_frame, text="添加版本号",
                         variable=self.rename_action, value="version",
                         command=self._on_action_change).pack(side="left", padx=(0, 12))
        ttk.Radiobutton(action_frame, text="添加指定字符",
                         variable=self.rename_action, value="custom",
                         command=self._on_action_change).pack(side="left")

        self.param_frame = tk.Frame(rule_card, bg=CARD_COLOR)
        self.param_frame.pack(fill="x", pady=(0, 4))
        self._build_param_controls()

        self.pos_frame = tk.Frame(rule_card, bg=CARD_COLOR)
        self.pos_frame.pack(fill="x", pady=(0, 4))
        tk.Label(self.pos_frame, text="添加位置：",
                 font=("Microsoft YaHei", 10),
                 bg=CARD_COLOR, fg=TEXT_PRIMARY).pack(side="left")
        ttk.Radiobutton(self.pos_frame, text="文件名前",
                         variable=self.add_position, value="prefix",
                         command=self._auto_preview).pack(side="left", padx=(8, 12))
        ttk.Radiobutton(self.pos_frame, text="文件名后",
                         variable=self.add_position, value="suffix",
                         command=self._auto_preview).pack(side="left")

        path_card = ttk.LabelFrame(container, text=" 文件夹路径 ",
                                    style="Card.TLabelframe", padding=10)
        path_card.pack(fill="x", pady=(0, 8))
        path_inner = tk.Frame(path_card, bg=CARD_COLOR)
        path_inner.pack(fill="x")
        tk.Entry(path_inner, textvariable=self.folder_path,
                  font=("Microsoft YaHei", 10), relief="solid", bd=1).pack(side="left", fill="x", expand=True, padx=(0, 8), pady=6)
        ttk.Button(path_inner, text="浏览...",
                    style="Accent.TButton",
                    command=self._browse_folder, width=8).pack(side="right", pady=6)

        action_row = tk.Frame(container, bg=BG_COLOR)
        action_row.pack(fill="x", pady=(0, 8))
        action_row.columnconfigure(0, weight=1)
        action_row.columnconfigure(1, weight=1)
        action_row.columnconfigure(2, weight=1)

        self.start_btn = ttk.Button(action_row, text="▶  开始批量重命名",
                                     style="Success.TButton",
                                     command=self._start_rename)
        self.start_btn.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        preview_btn = ttk.Button(action_row, text="👁  预览",
                                  style="Ghost.TButton",
                                  command=self._preview_rename)
        preview_btn.grid(row=0, column=1, sticky="ew", padx=6)
        self.undo_btn = ttk.Button(action_row, text="↩  后悔，回退",
                                    style="Danger.TButton",
                                    command=self._undo_rename, state="disabled")
        self.undo_btn.grid(row=0, column=2, sticky="ew", padx=(6, 0))

        progress_card = ttk.LabelFrame(container, text=" 处理进度 ",
                                        style="Card.TLabelframe", padding=10)
        progress_card.pack(fill="x", pady=(0, 8))
        self.progress_bar = ttk.Progressbar(progress_card, variable=self.progress_var,
                                             maximum=100, style="Horizontal.TProgressbar")
        self.progress_bar.pack(fill="x", pady=(0, 6))
        tk.Label(progress_card, textvariable=self.status_text,
                  font=("Microsoft YaHei", 9),
                  fg=TEXT_SECONDARY, bg=CARD_COLOR).pack(anchor="w")

        log_card = ttk.LabelFrame(container, text=" 预览 / 日志 ",
                                    style="Card.TLabelframe", padding=10)
        log_card.pack(fill="both", expand=True)
        self.log_text = tk.Text(log_card, height=15, font=("Consolas", 9),
                                 state="disabled", relief="flat",
                                 bg="#FAFAFA", fg=TEXT_PRIMARY,
                                 selectbackground=SUCCESS, selectforeground="white",
                                 wrap="word", padx=8, pady=6)
        self.log_text.pack(fill="both", expand=True)

    def _log(self, message):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{ts}] {message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self.root.update_idletasks()

    def _clear_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    def _go_main(self):
        self.root.destroy()
        main = tk.Tk()
        MainApp(main)
        main.mainloop()

    def _build_param_controls(self):
        for w in self.param_frame.winfo_children():
            w.destroy()
        action = self.rename_action.get()
        if action == "date":
            tk.Label(self.param_frame, text="日期格式：",
                     font=("Microsoft YaHei", 10),
                     bg=CARD_COLOR, fg=TEXT_PRIMARY).pack(side="left")
            formats = ["%Y%m%d", "%Y-%m-%d", "%Y%m%d_%H%M", "cn_date"]
            format_combo = ttk.Combobox(self.param_frame, values=formats,
                                         state="readonly", width=15,
                                         font=("Microsoft YaHei", 9))
            format_combo.set(self.date_format.get())
            format_combo.pack(side="left", padx=(4, 0))
            format_combo.bind("<<ComboboxSelected>>",
                               lambda e: self._on_date_format_change(format_combo))
        elif action == "version":
            tk.Label(self.param_frame, text="版本号文字：",
                     font=("Microsoft YaHei", 10),
                     bg=CARD_COLOR, fg=TEXT_PRIMARY).pack(side="left")
            ver_entry = tk.Entry(self.param_frame, textvariable=self.version_text,
                                  font=("Microsoft YaHei", 10), width=15, relief="solid", bd=1)
            ver_entry.pack(side="left", padx=(4, 0))
            ver_entry.bind("<KeyRelease>", lambda e: self._auto_preview())
        elif action == "custom":
            tk.Label(self.param_frame, text="字符内容：",
                     font=("Microsoft YaHei", 10),
                     bg=CARD_COLOR, fg=TEXT_PRIMARY).pack(side="left")
            char_entry = tk.Entry(self.param_frame, textvariable=self.custom_chars,
                                   font=("Microsoft YaHei", 10), width=10, relief="solid", bd=1)
            char_entry.pack(side="left", padx=(4, 0))
            char_entry.bind("<KeyRelease>", lambda e: self._auto_preview())
            tk.Label(self.param_frame, text=" 重复",
                     font=("Microsoft YaHei", 10),
                     bg=CARD_COLOR, fg=TEXT_PRIMARY).pack(side="left", padx=(4, 0))
            count_combo = ttk.Combobox(self.param_frame,
                                        values=[str(i) for i in range(1, 101)],
                                        width=4, state="readonly",
                                        font=("Microsoft YaHei", 9))
            count_combo.set(str(self.custom_count.get()))
            count_combo.pack(side="left", padx=(4, 0))
            count_combo.bind("<<ComboboxSelected>>",
                             lambda e: self._on_count_change(count_combo))
            tk.Label(self.param_frame, text=" 次",
                     font=("Microsoft YaHei", 10),
                     bg=CARD_COLOR, fg=TEXT_PRIMARY).pack(side="left")

    def _on_action_change(self):
        self._build_param_controls()
        self._auto_preview()

    def _on_date_format_change(self, combo):
        self.date_format.set(combo.get())
        self._auto_preview()

    def _on_count_change(self, combo):
        self.custom_count.set(int(combo.get()))
        self._auto_preview()

    def _browse_folder(self):
        folder = filedialog.askdirectory(title="请选择要处理的文件夹")
        if folder:
            self.folder_path.set(folder)
            self._auto_preview()

    def _get_files(self, path):
        try:
            items = os.listdir(path)
            files = [i for i in items if os.path.isfile(os.path.join(path, i))]
            return files
        except Exception as e:
            messagebox.showerror("错误", f"无法读取文件夹：{e}")
            return []

    def _compute_new_name(self, filename):
        name, ext = os.path.splitext(filename)
        action = self.rename_action.get()
        pos = self.add_position.get()

        if action == "date":
            fmt = self.date_format.get()
            now = datetime.now()
            if fmt == "cn_date":
                date_str = now.strftime("%Y年%m月%d日")
            else:
                date_str = now.strftime(fmt)
            if pos == "prefix":
                return f"{date_str}_{name}{ext}"
            else:
                return f"{name}_{date_str}{ext}"
        elif action == "version":
            ver = self.version_text.get()
            if pos == "prefix":
                return f"{ver}_{name}{ext}"
            else:
                return f"{name}_{ver}{ext}"
        elif action == "custom":
            chars = self.custom_chars.get()
            count = self.custom_count.get()
            repeated = chars * count
            if pos == "prefix":
                return f"{repeated}{name}{ext}"
            else:
                return f"{name}{repeated}{ext}"
        return None

    def _auto_preview(self):
        folder = self.folder_path.get().strip()
        if not folder or not os.path.isdir(folder):
            return
        self._clear_log()
        files = self._get_files(folder)
        if not files:
            self._log("[预览] 当前文件夹下没有文件。")
            return
        rename_list = []
        conflict_list = []
        for f in files:
            new_name = self._compute_new_name(f)
            if new_name:
                new_path = os.path.join(folder, new_name)
                if os.path.exists(new_path) and new_path != os.path.join(folder, f):
                    conflict_list.append((f, new_name))
                rename_list.append((f, new_name))
        self._log(f"[预览] 共 {len(files)} 个文件：")
        self._log(f"  待重命名：{len(rename_list)} 个")
        self._log("")
        for old, new in rename_list[:20]:
            flag = " ⚠冲突" if (old, new) in conflict_list else ""
            self._log(f"  {old}  →  {new}{flag}")
        if len(rename_list) > 20:
            self._log(f"  ... 仅显示前20个，共 {len(rename_list)} 个")
        self._log(f"[预览] 可重命名 {len(rename_list) - len(conflict_list)} 个")

    def _preview_rename(self):
        folder = self.folder_path.get().strip()
        if not folder:
            messagebox.showwarning("警告", "请先选择一个文件夹！")
            return
        self._auto_preview()

    def _start_rename(self):
        folder = self.folder_path.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("警告", "请先选择有效的文件夹！")
            return
        files = self._get_files(folder)
        if not files:
            messagebox.showinfo("提示", "没有文件。")
            return
        self._clear_log()
        self.rename_history = []
        self.is_processing = True
        self.start_btn.configure(state="disabled")
        self.undo_btn.configure(state="disabled")
        threading.Thread(target=self._do_rename,
                         args=(folder, files), daemon=True).start()

    def _do_rename(self, folder, files):
        success_count = fail_count = conflict_count = 0
        total = len(files)
        start_time = datetime.now()
        self.progress_var.set(0)
        self.status_text.set("正在处理...")

        for i, fname in enumerate(files):
            old_path = os.path.join(folder, fname)
            new_name = self._compute_new_name(fname)
            if not new_name:
                fail_count += 1
                continue
            new_path = os.path.join(folder, new_name)
            if os.path.exists(new_path) and new_path != old_path:
                conflict_count += 1
                fail_count += 1
            else:
                try:
                    os.rename(old_path, new_path)
                    success_count += 1
                    self.rename_history.append((folder, fname, new_name))
                except Exception as e:
                    fail_count += 1
            progress = (i + 1) / total * 100
            self.progress_var.set(progress)
            self.status_text.set(f"处理中... {i+1}/{total}")
            self.root.update_idletasks()

        duration = (datetime.now() - start_time).total_seconds()
        summary = f"完成！成功 {success_count}，失败 {fail_count}"
        self.status_text.set(summary)
        self.progress_var.set(100)
        self.start_btn.configure(state="normal")
        self.is_processing = False
        if success_count > 0 and self.rename_history:
            self.undo_btn.configure(state="normal")
        messagebox.showinfo("完成",
            f"成功：{success_count}\n失败：{fail_count}\n耗时：{duration:.2f}s")

    def _undo_rename(self):
        if self.is_processing:
            return
        if not self.rename_history:
            messagebox.showinfo("提示", "没有可回退的操作。")
            return
        count = len(self.rename_history)
        if not messagebox.askyesno("确认",
            f"回退最近的 {count} 个重命名操作？"):
            return
        self.start_btn.configure(state="disabled")
        self.undo_btn.configure(state="disabled")
        undo_history = list(reversed(self.rename_history))
        self.rename_history = []
        self.is_processing = True
        success_count = fail_count = 0
        total = len(undo_history)
        self.progress_var.set(0)
        self.status_text.set("正在回退...")
        for i, (f, orig, curr) in enumerate(undo_history):
            curr_path = os.path.join(f, curr)
            orig_path = os.path.join(f, orig)
            if not os.path.exists(curr_path):
                fail_count += 1
            elif os.path.exists(orig_path) and orig_path != curr_path:
                fail_count += 1
            else:
                try:
                    os.rename(curr_path, orig_path)
                    success_count += 1
                except Exception:
                    fail_count += 1
            progress = (i + 1) / total * 100
            self.progress_var.set(progress)
            self.status_text.set(f"回退中... {i+1}/{total}")
            self.root.update_idletasks()
        summary = f"回退完成！成功 {success_count}，失败 {fail_count}"
        self.status_text.set(summary)
        self.progress_var.set(100)
        self.start_btn.configure(state="normal")
        self.is_processing = False
        messagebox.showinfo("回退完成", summary)


class SpecUploadApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"规约上传数据表准备 v{APP_VERSION}")
        self.root.geometry("800x720")
        self.root.configure(bg=BG_COLOR)
        self.root.resizable(False, False)

        self.excel_path = tk.StringVar()
        self.folder_path = tk.StringVar()
        self.left_compare_column = tk.StringVar(
            value=SPEC_DEFAULT_LEFT_COLUMN
        )
        self.right_compare_column = tk.StringVar(
            value=SPEC_DEFAULT_RIGHT_COLUMN
        )
        self.status_text = tk.StringVar(value="请选择 WPS 表格文件和目标文件夹")
        self.progress_var = tk.DoubleVar(value=0.0)

        self.generated_file = None
        self.is_processing = False

        self._setup_styles()
        self._build_ui()

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Card.TLabelframe",
                        background=CARD_COLOR, foreground=TEXT_PRIMARY,
                        bordercolor=BORDER)
        style.configure("Card.TLabelframe.Label",
                        font=("Microsoft YaHei", 10, "bold"),
                        foreground="#FF9800", background=CARD_COLOR)
        style.configure("Accent.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground="white", background=ACCENT,
                        borderwidth=0, padding=(12, 8))
        style.map("Accent.TButton", background=[("active", "#1976D2")])
        style.configure("Success.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground="white", background=SUCCESS,
                        borderwidth=0, padding=(12, 8))
        style.map("Success.TButton", background=[("active", "#388E3C")])
        style.configure("Warning.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground="white", background="#FF9800",
                        borderwidth=0, padding=(12, 8))
        style.map("Warning.TButton", background=[("active", "#F57C00")])
        style.configure("Ghost.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground=TEXT_PRIMARY, background=CARD_COLOR,
                        borderwidth=1, relief="solid", padding=(12, 8))
        style.map("Ghost.TButton", background=[("active", "#F5F5F5")])
        style.configure("Horizontal.TProgressbar",
                        troughcolor=BORDER, background="#FF9800",
                        thickness=8, borderwidth=0)

    def _build_ui(self):
        container = tk.Frame(self.root, bg=BG_COLOR)
        container.pack(fill="both", expand=True, padx=16, pady=12)

        header = tk.Frame(container, bg=BG_COLOR)
        header.pack(fill="x", pady=(0, 8))
        ttk.Label(header, text="📋 规约上传数据表准备 — 上传第三方填写审定信息",
                  font=("Microsoft YaHei", 14, "bold"),
                  foreground="#FF9800", background=BG_COLOR).pack(side="left")
        back_btn = tk.Label(header, text="← 返回主菜单",
                            font=("Microsoft YaHei", 9),
                            fg=TEXT_SECONDARY, bg=BG_COLOR,
                            cursor="hand2")
        back_btn.pack(side="right")
        back_btn.bind("<Button-1>", lambda e: self._go_main())

        sep = tk.Frame(container, bg=BORDER, height=1)
        sep.pack(fill="x", pady=(0, 10))

        file_card = ttk.LabelFrame(container, text=" 文件选择 ",
                                    style="Card.TLabelframe", padding=10)
        file_card.pack(fill="x", pady=(0, 8))
        file_inner = tk.Frame(file_card, bg=CARD_COLOR)
        file_inner.pack(fill="x")

        tk.Label(file_inner, text="WPS表格：",
                 font=("Microsoft YaHei", 10, "bold"),
                 bg=CARD_COLOR, fg=TEXT_PRIMARY).grid(row=0, column=0, sticky="w")
        tk.Entry(file_inner, textvariable=self.excel_path,
                  font=("Microsoft YaHei", 10), relief="solid", bd=1,
                  width=55).grid(row=0, column=1, sticky="we", padx=(4, 4), pady=4)
        ttk.Button(file_inner, text="浏览...",
                    style="Accent.TButton",
                    command=self._browse_excel, width=8).grid(row=0, column=2, pady=4)

        tk.Label(file_inner, text="目标文件夹：",
                 font=("Microsoft YaHei", 10, "bold"),
                 bg=CARD_COLOR, fg=TEXT_PRIMARY).grid(row=1, column=0, sticky="w", pady=(6, 0))
        tk.Entry(file_inner, textvariable=self.folder_path,
                  font=("Microsoft YaHei", 10), relief="solid", bd=1,
                  width=55).grid(row=1, column=1, sticky="we", padx=(4, 4), pady=(6, 0))
        ttk.Button(file_inner, text="浏览...",
                    style="Accent.TButton",
                    command=self._browse_folder, width=8).grid(row=1, column=2, pady=(6, 0))

        file_inner.columnconfigure(1, weight=1)

        rule_card = ttk.LabelFrame(container, text=" 比对规则 ",
                                    style="Card.TLabelframe", padding=10)
        rule_card.pack(fill="x", pady=(0, 8))
        rule_inner = tk.Frame(rule_card, bg=CARD_COLOR)
        rule_inner.pack(fill="x")

        tk.Label(rule_inner, text="文件名格式：",
                 font=("Microsoft YaHei", 10, "bold"),
                 bg=CARD_COLOR, fg=TEXT_PRIMARY).grid(row=0, column=0, sticky="w")
        tk.Label(rule_inner, text="横杠左边 - 横杠右边",
                 font=("Microsoft YaHei", 11),
                 bg=CARD_COLOR, fg=ACCENT).grid(row=0, column=1, sticky="w", padx=(4, 12))
        tk.Label(rule_inner, text="任一侧匹配；使用现有 K/L，不新增列",
                 font=("Microsoft YaHei", 9),
                 bg=CARD_COLOR, fg=TEXT_SECONDARY).grid(
                     row=0, column=2, columnspan=4, sticky="w"
                 )

        tk.Label(rule_inner, text="横杠左边 → 表格",
                 font=("Microsoft YaHei", 9),
                 bg=CARD_COLOR, fg=TEXT_SECONDARY).grid(
                     row=1, column=0, sticky="w", pady=(8, 0)
                 )
        ttk.Combobox(
            rule_inner,
            textvariable=self.left_compare_column,
            values=SPEC_COLUMN_CHOICES,
            width=6,
        ).grid(row=1, column=1, sticky="w", padx=(4, 16), pady=(8, 0))
        tk.Label(rule_inner, text="横杠右边 → 表格",
                 font=("Microsoft YaHei", 9),
                 bg=CARD_COLOR, fg=TEXT_SECONDARY).grid(
                     row=1, column=2, sticky="w", pady=(8, 0)
                 )
        ttk.Combobox(
            rule_inner,
            textvariable=self.right_compare_column,
            values=SPEC_COLUMN_CHOICES,
            width=6,
        ).grid(row=1, column=3, sticky="w", padx=(4, 16), pady=(8, 0))
        tk.Label(rule_inner, text="默认 S / A；可直接输入 A-XFD",
                 font=("Microsoft YaHei", 9),
                 bg=CARD_COLOR, fg=TEXT_MUTED).grid(
                     row=1, column=4, columnspan=2, sticky="w", pady=(8, 0)
                 )

        action_row = tk.Frame(container, bg=BG_COLOR)
        action_row.pack(fill="x", pady=(0, 8))
        action_row.columnconfigure(0, weight=1)
        action_row.columnconfigure(1, weight=1)
        action_row.columnconfigure(2, weight=1)

        self.start_btn = ttk.Button(action_row, text="▶  开始比对并生成",
                                     style="Warning.TButton",
                                     command=self._start_compare)
        self.start_btn.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        preview_btn = ttk.Button(action_row, text="👁  预览比对",
                                  style="Ghost.TButton",
                                  command=self._preview_compare)
        preview_btn.grid(row=0, column=1, sticky="ew", padx=6)
        self.open_btn = ttk.Button(action_row, text="📂  打开生成的表格",
                                    style="Success.TButton",
                                    command=self._open_generated, state="disabled")
        self.open_btn.grid(row=0, column=2, sticky="ew", padx=(6, 0))

        progress_card = ttk.LabelFrame(container, text=" 处理进度 ",
                                        style="Card.TLabelframe", padding=10)
        progress_card.pack(fill="x", pady=(0, 8))
        self.progress_bar = ttk.Progressbar(progress_card, variable=self.progress_var,
                                             maximum=100, style="Horizontal.TProgressbar")
        self.progress_bar.pack(fill="x", pady=(0, 6))
        tk.Label(progress_card, textvariable=self.status_text,
                  font=("Microsoft YaHei", 9),
                  fg=TEXT_SECONDARY, bg=CARD_COLOR).pack(anchor="w")

        log_card = ttk.LabelFrame(container, text=" 日志 / 报告 ",
                                    style="Card.TLabelframe", padding=10)
        log_card.pack(fill="both", expand=True)
        self.log_text = tk.Text(log_card, height=12, font=("Consolas", 9),
                                 state="disabled", relief="flat",
                                 bg="#FAFAFA", fg=TEXT_PRIMARY,
                                 selectbackground="#FF9800", selectforeground="white",
                                 wrap="word", padx=8, pady=6)
        self.log_text.pack(fill="both", expand=True)

    def _log(self, message):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{ts}] {message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self.root.update_idletasks()

    def _clear_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    def _go_main(self):
        self.root.destroy()
        main = tk.Tk()
        MainApp(main)
        main.mainloop()

    def _browse_excel(self):
        file_path = filedialog.askopenfilename(
            title="选择 WPS 表格文件",
            filetypes=[("表格文件", "*.xlsx *.xls *.et"), ("所有文件", "*.*")])
        if file_path:
            self.excel_path.set(file_path)
            self.status_text.set(f"已选择表格：{os.path.basename(file_path)}")

    def _browse_folder(self):
        folder = filedialog.askdirectory(title="选择目标文件夹")
        if folder:
            self.folder_path.set(folder)
            self.status_text.set(f"已选择文件夹：{folder}")

    def _read_excel_data(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=1, values_only=False):
            row_data = []
            for cell in row:
                row_data.append(str(cell.value) if cell.value is not None else "")
            data.append(row_data)
        wb.close()
        return data, ws.title

    def _scan_folder_files(self, folder_path):
        files_info = []
        for fname in os.listdir(folder_path):
            fpath = os.path.join(folder_path, fname)
            if os.path.isfile(fpath):
                name_without_ext = os.path.splitext(fname)[0]
                if "-" in name_without_ext:
                    parts = name_without_ext.split("-", 1)
                    left = parts[0].strip()
                    right = parts[1].strip()
                else:
                    left = name_without_ext.strip()
                    right = ""
                files_info.append({
                    "filename": fname,
                    "left": left,
                    "right": right,
                    "filepath": fpath
                })
        return files_info

    def _get_compare_columns(self):
        compare_columns = _normalize_spec_compare_columns(
            self.left_compare_column.get(),
            self.right_compare_column.get(),
        )
        self.left_compare_column.set(compare_columns[0])
        self.right_compare_column.set(compare_columns[2])
        return compare_columns

    def _preview_compare(self):
        excel_file = self.excel_path.get().strip()
        folder = self.folder_path.get().strip()

        if not excel_file or not os.path.exists(excel_file):
            messagebox.showwarning("提示", "请先选择有效的 WPS 表格文件")
            return
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("提示", "请先选择有效的目标文件夹")
            return

        self._clear_log()
        self._log("开始预览比对...")

        try:
            compare_columns = self._get_compare_columns()
            left_label, left_index, right_label, right_index = compare_columns
            excel_data, sheet_name = self._read_excel_data(excel_file)
            _validate_spec_compare_columns(excel_data, compare_columns)
            _validate_spec_status_columns(excel_data)
            self._log(f"读取表格：{os.path.basename(excel_file)} (工作表: {sheet_name})")
            self._log(f"表格共 {len(excel_data)} 行")
            self._log(
                f"比对规则：横杠左边 → {left_label}列；"
                f"横杠右边 → {right_label}列"
            )

            folder_files = self._scan_folder_files(folder)
            self._log(f"扫描文件夹：共 {len(folder_files)} 个文件")

            matches = 0
            duplicate_matches = 0
            for i, row in enumerate(
                    excel_data[SPEC_DATA_START_ROW - 1:],
                    start=SPEC_DATA_START_ROW):
                left_value = _spec_row_value(row, left_index)
                right_value = _spec_row_value(row, right_index)
                if not left_value and not right_value:
                    continue
                for fi in folder_files:
                    if (
                        left_value and fi["left"] == left_value
                    ) or (
                        right_value and fi["right"] == right_value
                    ):
                        matches += 1
                        checkbox_value = row[SPEC_CHECKBOX_COLUMN - 1]
                        upload_status = row[SPEC_UPLOAD_STATUS_COLUMN - 1]
                        is_duplicate = _spec_status_is_marked(
                            checkbox_value, upload_status
                        )
                        if is_duplicate:
                            duplicate_matches += 1
                        self._log(
                            f"  {'重复' if is_duplicate else '新增'}匹配："
                            f"表格第{i}行 "
                            f"[{left_label}列:{left_value} | "
                            f"{right_label}列:{right_value}] "
                            f"↔ [{fi['filename']}]"
                        )
                        break

            new_matches = matches - duplicate_matches
            self._log(
                f"预览完成：找到 {matches} 条匹配记录，"
                f"新增 {new_matches} 条，重复 {duplicate_matches} 条"
            )
            data_row_count = max(0, len(excel_data) - SPEC_HEADER_ROWS)
            self.status_text.set(
                f"预览：共 {data_row_count} 行数据，"
                f"{len(folder_files)} 个文件，新增 {new_matches} 条，"
                f"重复 {duplicate_matches} 条"
            )

        except Exception as e:
            self._log(f"错误：{e}")
            messagebox.showerror("预览失败", str(e))

    def _start_compare(self):
        excel_file = self.excel_path.get().strip()
        folder = self.folder_path.get().strip()

        if not excel_file or not os.path.exists(excel_file):
            messagebox.showwarning("提示", "请先选择有效的 WPS 表格文件")
            return
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("提示", "请先选择有效的目标文件夹")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror("依赖缺失", "缺少 openpyxl 库")
            return
        try:
            compare_columns = self._get_compare_columns()
        except ValueError as error:
            messagebox.showwarning("比对列无效", str(error))
            return

        self._clear_log()
        self.start_btn.configure(state="disabled")
        self.progress_var.set(0)
        self.is_processing = True

        threading.Thread(
            target=self._do_compare,
            args=(excel_file, folder, compare_columns),
            daemon=True,
        ).start()

    def _do_compare(self, excel_file, folder, compare_columns):
        try:
            left_label, left_index, right_label, right_index = compare_columns

            self.root.after(0, self._log, "==== 规约上传数据表准备：比对开始 ====")
            self.root.after(0, self._log, f"源表格：{excel_file}")
            self.root.after(0, self._log, f"目标文件夹：{folder}")
            self.root.after(
                0, self._log,
                f"比对规则：横杠左边 → {left_label}列；"
                f"横杠右边 → {right_label}列"
            )

            excel_data, sheet_name = self._read_excel_data(excel_file)
            _validate_spec_compare_columns(excel_data, compare_columns)
            _validate_spec_status_columns(excel_data)
            self.root.after(0, self._log, f"读取表格：{len(excel_data)} 行，工作表：{sheet_name}")

            folder_files = self._scan_folder_files(folder)
            self.root.after(0, self._log, f"扫描文件：{len(folder_files)} 个")

            base_name, ext = os.path.splitext(excel_file)
            output_path = f"{base_name}上传{ext}"
            shutil.copy2(excel_file, output_path)
            self.root.after(0, self._log, f"复制表格：{output_path}")

            self.root.after(0, self._log, "正在识别匹配行...")

            row_match_info = {}
            row_status_info = {}
            for i in range(SPEC_DATA_START_ROW, len(excel_data) + 1):
                row_data = excel_data[i - 1] if i - 1 < len(excel_data) else []
                left_value = _spec_row_value(row_data, left_index)
                right_value = _spec_row_value(row_data, right_index)
                checkbox_value = row_data[SPEC_CHECKBOX_COLUMN - 1]
                upload_status = row_data[SPEC_UPLOAD_STATUS_COLUMN - 1]
                matched = False
                for fi in folder_files:
                    if (
                        left_value and fi["left"] == left_value
                    ) or (
                        right_value and fi["right"] == right_value
                    ):
                        matched = True
                        break
                row_match_info[i] = matched
                row_status_info[i] = {
                    "marked": _spec_status_is_marked(
                        checkbox_value, upload_status
                    ),
                    "checkbox_value": checkbox_value,
                    "upload_status": upload_status,
                }

            matches = sum(1 for v in row_match_info.values() if v)
            duplicate_matches = sum(
                1 for row_number, matched in row_match_info.items()
                if matched and row_status_info[row_number]["marked"]
            )
            new_matches = matches - duplicate_matches
            total_data_rows = max(0, len(excel_data) - SPEC_HEADER_ROWS)
            self.root.after(
                0, self._log,
                f"匹配完成：共 {matches} 条，新增 {new_matches} 条，"
                f"重复 {duplicate_matches} 条，"
                f"{total_data_rows - matches} 条本次未匹配"
            )

            wb = openpyxl.load_workbook(output_path)
            ws = wb.active

            max_orig_col = ws.max_column
            if max_orig_col < SPEC_UPLOAD_STATUS_COLUMN:
                raise ValueError(
                    "表格中不存在现有 K/L 状态列，已停止处理且不会新增列"
                )
            matched_orig = [
                i for i in range(SPEC_DATA_START_ROW, len(excel_data) + 1)
                if row_match_info[i]
            ]
            unmatched_orig = [
                i for i in range(SPEC_DATA_START_ROW, len(excel_data) + 1)
                if not row_match_info[i]
            ]
            new_order = matched_orig + unmatched_orig

            data_start_row = SPEC_DATA_START_ROW
            data_end_row = len(excel_data)
            _reorder_worksheet_rows(
                ws, new_order, data_start_row, max_orig_col
            )
            self.root.after(0, self._log,
                f"重排序完成：前{SPEC_HEADER_ROWS}行格式完整保留，"
                f"待上传行从第{data_start_row}行开始置顶")

            col_checkbox = SPEC_CHECKBOX_COLUMN
            col_need_upload = SPEC_UPLOAD_STATUS_COLUMN
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin", color="FFE0E0E0"),
                right=Side(style="thin", color="FFE0E0E0"),
                top=Side(style="thin", color="FFE0E0E0"),
                bottom=Side(style="thin", color="FFE0E0E0")
            )

            k_col_letter = get_column_letter(col_checkbox)
            l_col_letter = get_column_letter(col_need_upload)
            status_rows = []
            checkbox_values = {}
            for target_row, source_row in enumerate(
                    new_order, start=data_start_row):
                status_info = row_status_info[source_row]
                check_cell = ws.cell(
                    row=target_row, column=col_checkbox
                )
                need_cell = ws.cell(
                    row=target_row, column=col_need_upload
                )

                if (
                    status_info["marked"]
                    and _is_spec_status_formula(need_cell.value)
                ):
                    need_cell.value = _spec_status_formula(target_row)

                if row_match_info[source_row]:
                    if status_info["marked"]:
                        need_cell.value = SPEC_DUPLICATE_STATUS
                    else:
                        check_cell.value = ""
                        need_cell.value = _spec_status_formula(target_row)
                        check_cell.alignment = center_align
                        need_cell.alignment = center_align
                        if check_cell.border == Border():
                            check_cell.border = thin_border
                        if need_cell.border == Border():
                            need_cell.border = thin_border

                if _spec_status_is_marked(
                        check_cell.value, need_cell.value):
                    status_rows.append(target_row)
                    checkbox_values[target_row] = check_cell.value

            self.root.after(0, self.progress_var.set, 70)
            self.root.after(0, self._log, "正在设置列宽和条件格式...")

            col_width_need = max(ws.column_dimensions[l_col_letter].width or 0, 12)
            col_width_check = max(ws.column_dimensions[k_col_letter].width or 0, 6)
            ws.column_dimensions[l_col_letter].width = col_width_need
            ws.column_dimensions[k_col_letter].width = col_width_check

            red_bold_font = Font(bold=True, size=10, color="FFFF0000")
            green_text_font = Font(bold=True, size=10, color="FF00B050")
            light_green_fill = PatternFill(start_color="FFEBF1DE", end_color="FFEBF1DE", fill_type="solid")

            if data_end_row >= data_start_row:
                # Replace status rules from earlier runs instead of stacking them.
                for cf in list(ws.conditional_formatting._cf_rules):
                    rules = ws.conditional_formatting._cf_rules[cf]
                    remaining_rules = [
                        rule for rule in rules
                        if not any(
                            f"{k_col_letter}{data_start_row}=TRUE" in formula
                            for formula in (rule.formula or [])
                        )
                    ]
                    if remaining_rules:
                        ws.conditional_formatting._cf_rules[cf] = (
                            remaining_rules
                        )
                    else:
                        del ws.conditional_formatting._cf_rules[cf]

                max_col = max_orig_col
                row_range = (
                    f"A{data_start_row}:"
                    f"{get_column_letter(max_col)}{data_end_row}"
                )
                row_fill_rule = FormulaRule(
                    formula=[
                        f'AND(${k_col_letter}{data_start_row}=TRUE,'
                        f'${l_col_letter}{data_start_row}'
                        f'<>"{SPEC_DUPLICATE_STATUS}")'
                    ],
                    fill=light_green_fill
                )
                ws.conditional_formatting.add(row_range, row_fill_rule)

                l_cf_range = (
                    f"{l_col_letter}{data_start_row}:"
                    f"{l_col_letter}{data_end_row}"
                )
                red_bold_rule = FormulaRule(
                    formula=[
                        f"OR(NOT({k_col_letter}{data_start_row}=TRUE),"
                        f'{l_col_letter}{data_start_row}'
                        f'="{SPEC_DUPLICATE_STATUS}")'
                    ],
                    font=red_bold_font
                )
                ws.conditional_formatting.add(l_cf_range, red_bold_rule)

                green_bold_rule = FormulaRule(
                    formula=[
                        f"AND({k_col_letter}{data_start_row}=TRUE,"
                        f'{l_col_letter}{data_start_row}'
                        f'<>"{SPEC_DUPLICATE_STATUS}")'
                    ],
                    font=green_text_font
                )
                ws.conditional_formatting.add(l_cf_range, green_bold_rule)

            _set_worksheet_columns_hidden(
                ws, col_checkbox, col_need_upload
            )
            wb.save(output_path)
            wb.close()

            self.root.after(0, self.progress_var.set, 85)
            self.root.after(0, self._log, "正在同步 K 列窗体复选框...")
            self._sync_form_checkboxes(
                output_path,
                status_rows,
                checkbox_values,
                col_checkbox,
                col_need_upload,
            )
            if not _workbook_columns_are_hidden(
                    output_path, col_checkbox, col_need_upload):
                raise ValueError("生成文件 K/L 列隐藏折叠校验失败")
            self.root.after(
                0, self._log, "[校验通过] K/L 列已隐藏折叠"
            )

            self.root.after(0, self.progress_var.set, 100)
            self.root.after(0, self._log, f"==== 比对完成 ====")
            self.root.after(
                0, self._log,
                f"匹配成功：{matches} 条（新增 {new_matches} 条，"
                f"重复 {duplicate_matches} 条）"
            )
            self.root.after(0, self._log, f"生成文件：{output_path}")
            self.root.after(0, self._log, f"待上传行已置顶（共 {matches} 条，排在表格顶部）")

            self.generated_file = output_path
            self.root.after(0, lambda: self.open_btn.configure(state="normal"))
            self.root.after(0, lambda: self.status_text.set(
                f"完成！新增 {new_matches} 条，重复 "
                f"{duplicate_matches} 条（已置顶），已生成："
                f"{os.path.basename(output_path)}"))

            self.root.after(0, lambda: messagebox.showinfo("完成",
                f"比对完成！\n\n匹配：{matches} 条\n"
                f"新增：{new_matches} 条\n重复：{duplicate_matches} 条\n"
                f"已置顶：{matches} 条待上传行\n生成文件：\n"
                f"{output_path}\n\n请用 WPS 打开编辑"))

        except Exception as e:
            import traceback
            traceback.print_exc()
            err_msg = str(e)
            self.root.after(0, self._log, f"错误：{err_msg}")
            self.root.after(0, lambda: messagebox.showerror("处理失败", err_msg))
        finally:
            self.is_processing = False
            self.root.after(0, lambda: self.start_btn.configure(state="normal"))

    def _open_generated(self):
        if self.generated_file and os.path.exists(self.generated_file):
            try:
                os.startfile(self.generated_file)
                self._log(f"已打开：{self.generated_file}")
            except Exception as e:
                messagebox.showerror("打开失败", f"无法打开文件：{e}")
        else:
            messagebox.showwarning("提示", "请先执行比对生成表格")

    def _sync_form_checkboxes(
            self, file_path, status_rows, checkbox_values,
            col_checkbox, col_need_upload):
        try:
            import win32com.client
        except ImportError:
            self.root.after(
                0, self._log,
                "⚠ 缺少 pywin32，跳过复选框同步"
            )
            return False

        excel_app = None
        workbook = None
        operation_error = None
        try:
            prog_ids = ["Excel.Application", "et.Application", "Kwps.Application"]
            last_err = None
            for prog_id in prog_ids:
                try:
                    excel_app = win32com.client.Dispatch(prog_id)
                    test_workbooks = excel_app.Workbooks
                    self.root.after(0, self._log, f"已连接 {prog_id}")
                    break
                except Exception as e:
                    last_err = e
                    excel_app = None

            if excel_app is None:
                raise RuntimeError(
                    f"无法启动 WPS/Excel：{last_err}"
                )

            excel_app.Visible = False
            excel_app.DisplayAlerts = False

            workbook = excel_app.Workbooks.Open(os.path.abspath(file_path))
            ws = workbook.Worksheets(1)

            col_letter = get_column_letter(col_checkbox)
            status_columns = ws.Columns(
                f"{get_column_letter(col_checkbox)}:"
                f"{get_column_letter(col_need_upload)}"
            )
            status_columns.Hidden = False

            removed_count = 0
            checkboxes = ws.CheckBoxes()
            for index in range(checkboxes.Count, 0, -1):
                checkbox = checkboxes(index)
                if (
                    checkbox.TopLeftCell.Column == col_checkbox
                    and checkbox.TopLeftCell.Row >= SPEC_DATA_START_ROW
                ):
                    checkbox.Delete()
                    removed_count += 1

            for row_num in status_rows:
                cell_addr = f"{col_letter}{row_num}"
                cell = ws.Range(cell_addr)

                cb = ws.CheckBoxes().Add(
                    Left=cell.Left + 2,
                    Top=cell.Top + 2,
                    Width=cell.Width - 4,
                    Height=cell.Height - 4
                )
                cb.Caption = ""
                original_value = checkbox_values.get(row_num)
                checked = (
                    original_value is True
                    or original_value == 1
                    or str(original_value).strip().upper() == "TRUE"
                )
                cb.Value = 1 if checked else -4146

                try:
                    cb.LinkedCell = cell_addr
                except Exception:
                    try:
                        cb.ControlFormat.LinkedCell = cell_addr
                    except Exception:
                        pass
                cell.Value = (
                    "" if original_value in (None, "")
                    else original_value
                )

            status_columns.Hidden = True

            workbook.Save()

        except Exception as e:
            import traceback
            traceback.print_exc()
            operation_error = e
        finally:
            try:
                if workbook:
                    workbook.Close(SaveChanges=False)
                if excel_app:
                    excel_app.Quit()
            except Exception:
                pass

        if operation_error is not None:
            self.root.after(
                0, self._log,
                f"⚠ 复选框同步失败：{operation_error}"
            )
            self.root.after(
                0, self._log,
                "K/L 列已在生成阶段完成隐藏折叠"
            )
            return False

        self.root.after(
            0, self._log,
            f"已清理 {removed_count} 个旧控件并重建 "
            f"{len(status_rows)} 个窗体复选框，"
            "K/L 列已隐藏并分组折叠"
        )
        return True


class AIChatApp:
    def __init__(self, root):
        try:
            self._init_impl(root)
        except Exception as e:
            import traceback
            traceback.print_exc()
            try:
                messagebox.showerror("初始化错误", f"AI助手初始化失败：\n{e}")
            except Exception:
                pass
            raise

    def _init_impl(self, root):
        self.root = root
        self.root.title(f"AI 智能助手 v{APP_VERSION}")
        self.root.geometry("860x760")
        self.root.configure(bg=BG_COLOR)
        self.root.resizable(False, False)

        self.api_key = tk.StringVar()
        self.current_provider = tk.StringVar(value="硅基流动")
        self.current_model = tk.StringVar()
        self.temperature = tk.DoubleVar(value=0.7)
        self.chat_history = []
        self.is_sending = False
        self._auto_save_job = None

        self._load_api_key()
        self._setup_styles()
        self._build_ui()
        self._update_model_list()

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Send.TButton",
                        font=("Microsoft YaHei", 10, "bold"),
                        foreground="white", background=SUCCESS,
                        borderwidth=0, padding=(16, 10))
        style.map("Send.TButton", background=[("active", "#388E3C")])

    def _build_ui(self):
        container = tk.Frame(self.root, bg=BG_COLOR)
        container.pack(fill="both", expand=True, padx=12, pady=10)

        header = tk.Frame(container, bg=BG_COLOR)
        header.pack(fill="x", pady=(0, 6))
        tk.Label(header, text="🤖 AI 智能助手",
                 font=("Microsoft YaHei", 16, "bold"),
                 fg=PURPLE, bg=BG_COLOR).pack(side="left")
        back_btn = tk.Label(header, text="← 返回主菜单",
                            font=("Microsoft YaHei", 9),
                            fg=TEXT_SECONDARY, bg=BG_COLOR,
                            cursor="hand2")
        back_btn.pack(side="right")
        back_btn.bind("<Button-1>", lambda e: self._go_main())

        config_card = tk.Frame(container, bg=CARD_COLOR,
                                highlightbackground=BORDER, highlightthickness=1)
        config_card.pack(fill="x", pady=(0, 6))
        cfg_inner = tk.Frame(config_card, bg=CARD_COLOR)
        cfg_inner.pack(fill="x", padx=10, pady=8)

        tk.Label(cfg_inner, text="接口：",
                 font=("Microsoft YaHei", 9),
                 bg=CARD_COLOR, fg=TEXT_PRIMARY).grid(row=0, column=0, sticky="w")
        provider_combo = ttk.Combobox(cfg_inner,
                                        values=list(AI_PROVIDERS.keys()),
                                        state="readonly", width=14,
                                        font=("Microsoft YaHei", 9))
        provider_combo.set(self.current_provider.get())
        provider_combo.grid(row=0, column=1, sticky="w", padx=(4, 12))
        provider_combo.bind("<<ComboboxSelected>>",
                             lambda e: self._on_provider_change(provider_combo))

        tk.Label(cfg_inner, text="模型：",
                 font=("Microsoft YaHei", 9),
                 bg=CARD_COLOR, fg=TEXT_PRIMARY).grid(row=0, column=2, sticky="w")
        self.model_combo = ttk.Combobox(cfg_inner, state="readonly", width=20,
                                         font=("Microsoft YaHei", 9))
        self.model_combo.grid(row=0, column=3, sticky="w", padx=(4, 12))
        self.model_combo.bind("<<ComboboxSelected>>",
                               lambda e: self._on_model_change())

        tk.Label(cfg_inner, text="温度：",
                 font=("Microsoft YaHei", 9),
                 bg=CARD_COLOR, fg=TEXT_PRIMARY).grid(row=0, column=4, sticky="w")
        temp_scale = tk.Scale(cfg_inner, from_=0.3, to=1.0, resolution=0.1,
                               orient="horizontal", variable=self.temperature,
                               bg=CARD_COLOR, fg=TEXT_PRIMARY,
                               highlightthickness=0, length=100, showvalue=True)
        temp_scale.grid(row=0, column=5, sticky="w", padx=(4, 0))

        tk.Label(cfg_inner, text="API Key：",
                 font=("Microsoft YaHei", 9),
                 bg=CARD_COLOR, fg=TEXT_PRIMARY).grid(row=1, column=0, sticky="w", pady=(6, 0))
        self.key_entry = tk.Entry(cfg_inner, textvariable=self.api_key,
                                   font=("Microsoft YaHei", 9), width=40,
                                   relief="solid", bd=1, show="*")
        self.key_entry.grid(row=1, column=1, columnspan=4, sticky="we", padx=(4, 4), pady=(6, 0))
        self.key_entry.bind("<FocusOut>", lambda e: self._save_api_key())
        self.key_entry.bind("<KeyRelease>", self._on_key_edit)

        save_key_btn = tk.Label(cfg_inner, text="💾 保存Key",
                                font=("Microsoft YaHei", 9),
                                fg=ACCENT, bg=CARD_COLOR, cursor="hand2")
        save_key_btn.grid(row=1, column=5, sticky="w", padx=(4, 0), pady=(6, 0))
        save_key_btn.bind("<Button-1>", lambda e: self._save_api_key_now())

        self.toggle_key_btn = tk.Label(cfg_inner, text="显示/隐藏",
                                        font=("Microsoft YaHei", 8),
                                        fg=ACCENT, bg=CARD_COLOR, cursor="hand2")
        self.toggle_key_btn.grid(row=1, column=6, sticky="w", padx=(4, 0), pady=(6, 0))
        self.toggle_key_btn.bind("<Button-1>", lambda e: self._toggle_key())

        chat_card = tk.Frame(container, bg=CARD_COLOR,
                              highlightbackground=BORDER, highlightthickness=1)
        chat_card.pack(fill="both", expand=True, pady=(0, 6))

        self.chat_display = tk.Text(chat_card, font=("Microsoft YaHei", 10),
                                     state="disabled", relief="flat",
                                     bg="#FAFAFA", fg=TEXT_PRIMARY,
                                     wrap="word", padx=10, pady=8,
                                     spacing1=4, spacing3=4)
        self.chat_display.pack(fill="both", expand=True, padx=4, pady=4)
        self._append_chat("system", "欢迎使用 AI 智能助手！请先配置 API Key，然后开始对话。")

        input_card = tk.Frame(container, bg=CARD_COLOR,
                               highlightbackground=BORDER, highlightthickness=1)
        input_card.pack(fill="x")

        input_inner = tk.Frame(input_card, bg=CARD_COLOR)
        input_inner.pack(fill="x", padx=8, pady=8)

        self.input_text = tk.Text(input_inner, font=("Microsoft YaHei", 10),
                                    height=3, relief="flat", bg="#F5F5F5",
                                    fg=TEXT_PRIMARY, wrap="word", padx=6, pady=4)
        self.input_text.pack(side="left", fill="both", expand=True, padx=(0, 8))
        self.input_text.bind("<Control-Return>", lambda e: self._send_message())

        send_btn = ttk.Button(input_inner, text="发送 ⏎",
                               style="Send.TButton",
                               command=self._send_message)
        send_btn.pack(side="right")

        bottom_bar = tk.Frame(container, bg=BG_COLOR)
        bottom_bar.pack(fill="x", pady=(4, 0))
        tk.Label(bottom_bar, text="Ctrl+Enter 发送",
                 font=("Microsoft YaHei", 8),
                 fg=TEXT_MUTED, bg=BG_COLOR).pack(side="left")
        clear_btn = tk.Label(bottom_bar, text="🗑 清空对话",
                               font=("Microsoft YaHei", 9),
                               fg=DANGER, bg=BG_COLOR, cursor="hand2")
        clear_btn.pack(side="right")
        clear_btn.bind("<Button-1>", lambda e: self._clear_chat())
        save_btn = tk.Label(bottom_bar, text="💾 保存对话",
                              font=("Microsoft YaHei", 9),
                              fg=ACCENT, bg=BG_COLOR, cursor="hand2")
        save_btn.pack(side="right", padx=(0, 12))
        save_btn.bind("<Button-1>", lambda e: self._save_chat())

    def _go_main(self):
        self.root.destroy()
        main = tk.Tk()
        MainApp(main)
        main.mainloop()

    def _load_api_key(self):
        provider = self.current_provider.get()
        key_file = os.path.join(os.path.expanduser("~"), f".ai_key_{provider}")
        try:
            if os.path.exists(key_file):
                with open(key_file, "r") as f:
                    encoded = f.read().strip()
                    if encoded:
                        self.api_key.set(base64.b64decode(encoded).decode("utf-8"))
        except Exception:
            pass

    def _save_api_key(self):
        provider = self.current_provider.get()
        key_file = os.path.join(os.path.expanduser("~"), f".ai_key_{provider}")
        try:
            key = self.api_key.get().strip()
            with open(key_file, "w") as f:
                if key:
                    f.write(base64.b64encode(key.encode("utf-8")).decode("utf-8"))
                else:
                    f.write("")
        except Exception:
            pass

    def _on_key_edit(self, event=None):
        if self._auto_save_job:
            self.root.after_cancel(self._auto_save_job)
        self._auto_save_job = self.root.after(800, self._save_api_key)

    def _save_api_key_now(self):
        self._save_api_key()
        self._append_chat("system", f"✅ API Key 已保存（{self.current_provider.get()}）")

    def _toggle_key(self):
        if self.key_entry.cget("show") == "*":
            self.key_entry.configure(show="")
        else:
            self.key_entry.configure(show="*")

    def _on_provider_change(self, combo):
        self._save_api_key()
        self.current_provider.set(combo.get())
        self.api_key.set("")
        self._load_api_key()
        self._update_model_list()

    def _on_model_change(self):
        self.current_model.set(self.model_combo.get())

    def _update_model_list(self):
        provider = self.current_provider.get()
        config = AI_PROVIDERS.get(provider, {})
        models = config.get("models", [])
        self.model_combo["values"] = models
        default = config.get("default_model", models[0] if models else "")
        self.model_combo.set(default)
        self.current_model.set(default)

    def _append_chat(self, role, text):
        self.chat_display.configure(state="normal")
        if role == "user":
            self.chat_display.insert("end", f"\n🧑 我：\n{text}\n", "user")
        elif role == "assistant":
            self.chat_display.insert("end", f"\n🤖 AI：\n{text}\n", "assistant")
        else:
            self.chat_display.insert("end", f"\n📢 {text}\n", "system")
        self.chat_display.see("end")
        self.chat_display.configure(state="disabled")

    def _send_message(self):
        if self.is_sending:
            return
        message = self.input_text.get("1.0", "end").strip()
        if not message:
            return
        key = self.api_key.get().strip()
        if not key:
            messagebox.showwarning("警告", "请先填写 API Key！")
            return
        self._save_api_key()
        model = self.current_model.get()
        if not model:
            messagebox.showwarning("警告", "请选择模型！")
            return

        self.input_text.delete("1.0", "end")
        self._append_chat("user", message)
        self.chat_history.append({"role": "user", "content": message})

        self.is_sending = True
        self._thinking_start_pos = self.chat_display.index("end-1c")
        self._append_chat("system", "正在思考...")

        provider = self.current_provider.get()
        model = self.current_model.get()
        temperature = self.temperature.get()
        history_snapshot = list(self.chat_history)

        threading.Thread(target=self._call_api,
                         args=(key, provider, model, temperature, history_snapshot),
                         daemon=True).start()

    def _call_api(self, key, provider, model, temperature, history):
        base_url = AI_PROVIDERS[provider]["base_url"]
        url = f"{base_url}/chat/completions"

        payload = {
            "model": model,
            "messages": history[-20:],
            "temperature": temperature,
        }
        headers = {
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
        }

        try:
            resp = requests.post(url, json=payload, headers=headers, timeout=30)
            if resp.status_code == 401:
                raise Exception("API Key 无效或已过期")
            elif resp.status_code == 429:
                raise Exception("请求过于频繁，请稍后再试")
            elif resp.status_code != 200:
                raise Exception(f"API 返回错误：{resp.status_code}")

            data = resp.json()
            reply = data["choices"][0]["message"]["content"]
            usage = data.get("usage", {})
            prompt_tokens = usage.get("prompt_tokens", 0)
            completion_tokens = usage.get("completion_tokens", 0)
            total_tokens = usage.get("total_tokens", 0)

            self.chat_history.append({"role": "assistant", "content": reply})

            def update_ui():
                self.chat_display.configure(state="normal")
                self.chat_display.delete(self._thinking_start_pos, "end")
                self.chat_display.configure(state="disabled")
                self._append_chat("assistant", reply)
                self._append_chat("system",
                    f"Token 用量：输入 {prompt_tokens} / 输出 {completion_tokens} / 合计 {total_tokens}")

            self.root.after(0, update_ui)

        except requests.exceptions.Timeout:
            self.root.after(0, self._remove_thinking_and_report, "⚠ 请求超时，请检查网络连接")
        except requests.exceptions.ConnectionError:
            self.root.after(0, self._remove_thinking_and_report, "⚠ 网络连接失败")
        except Exception as e:
            self.root.after(0, self._remove_thinking_and_report, f"⚠ 错误：{e}")
        finally:
            self.is_sending = False

    def _remove_thinking_and_report(self, msg):
        self.chat_display.configure(state="normal")
        self.chat_display.delete(self._thinking_start_pos, "end")
        self.chat_display.configure(state="disabled")
        self._append_chat("system", msg)

    def _clear_chat(self):
        self.chat_history = []
        self.chat_display.configure(state="normal")
        self.chat_display.delete("1.0", "end")
        self.chat_display.configure(state="disabled")
        self._append_chat("system", "对话已清空")

    def _save_chat(self):
        if not self.chat_history:
            messagebox.showinfo("提示", "没有对话可保存")
            return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt")],
            initialfile=f"AI对话_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        if file_path:
            try:
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write(f"AI 智能助手对话记录\n")
                    f.write(f"保存时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"接口：{self.current_provider.get()}  模型：{self.current_model.get()}\n")
                    f.write("=" * 50 + "\n\n")
                    for msg in self.chat_history:
                        role = "我" if msg["role"] == "user" else "AI"
                        f.write(f"[{role}]\n{msg['content']}\n\n")
                messagebox.showinfo("保存成功", f"对话已保存到：\n{file_path}")
            except Exception as e:
                messagebox.showerror("保存失败", str(e))


if __name__ == "__main__":
    import traceback as _tb

    def _global_handler(exc_type, exc_value, exc_tb):
        _tb.print_exception(exc_type, exc_value, exc_tb)
        try:
            messagebox.showerror("程序错误",
                f"发生错误：\n{exc_type.__name__}: {exc_value}")
        except Exception:
            pass

    import sys as _sys
    _sys.excepthook = _global_handler

    try:
        root = tk.Tk()
        app = MainApp(root)
        root.mainloop()
    except Exception as e:
        _tb.print_exc()
        try:
            messagebox.showerror("启动错误", f"程序启动失败：\n{e}")
        except Exception:
            pass
